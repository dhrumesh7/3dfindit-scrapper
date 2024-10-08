from tracemalloc import start
import requests
import json
import csv
import time
import shutil
import os
import string
from sanitize_filename import sanitize
import glob
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup

# Initialize Selenium WebDriver with options
options = Options()
options.add_argument("--headless")  # Run in headless mode (without opening a browser window)
options.add_argument("--disable-gpu")  # Disable GPU for better compatibility
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.set_window_size(1200, 2000)

# List to keep track of all processed product links
all_links = []

# Define folders for storing images, data, and JSON files
PICFOLDER = 'pictures'
DATAFOLDER = 'data'
JSONFOLDER = 'json'

# Get the current working directory
cwd = os.getcwd()

# Define paths for pictures, data, and JSON files
picpath = os.path.join(cwd, PICFOLDER)
datapath = os.path.join(cwd, DATAFOLDER)
jsonpath = os.path.join(cwd, JSONFOLDER)

# Create directories if they do not exist
if not os.path.isdir(picpath):
    os.mkdir(picpath)

if not os.path.isdir(datapath):
    os.mkdir(datapath)

if not os.path.isdir(jsonpath):
    os.mkdir(jsonpath)

# Function to download images associated with part numbers
def download_image(url, partnumber):
    try:
        response = requests.get(url)
        picfile = os.path.join(picpath, sanitize(partnumber) + '.jpg')
        if not os.path.isfile(picfile):  # Avoid downloading if the file already exists
            with open(picfile, 'wb') as out_file:
                out_file.write(response.content)
    except: 
        pass  # Ignore any errors during image download

def savehtml(filename, html):
    f = open(filename, "w", encoding="utf8")
    f.write(html)
    f.close() 

# Function to get response from url with retry mechanisam
def get_url(url):
    tries = 1
    returnResponse = dict()
    errorMessage = None

    while tries <= 3:  # Try a maximum of 3 times
        try:
            print('try', tries)
            driver.get(url)
            time.sleep(5)

            # Try to find the scrollable div
            scrollable_div = driver.find_element(By.CLASS_NAME, 'ReactVirtualized__List')

            last_height = driver.execute_script("return arguments[0].scrollHeight", scrollable_div)
            while True:
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scrollable_div)
                time.sleep(2)  # Wait for new content to load
                new_height = driver.execute_script("return arguments[0].scrollHeight", scrollable_div)
                if new_height == last_height:
                    break  # Break loop when no more content is being loaded
                last_height = new_height

            # If the scrolling and loading process was successful, return the driver
            returnResponse['type'] = 'Success'
            returnResponse['data'] = driver
            return returnResponse

        except Exception as e:
            if isinstance(e, NoSuchElementException):
                errorMessage = 'NoSuchElementException'
                print("Caught NoSuchElementException:", e)
            time.sleep(3)  # Wait before retrying

        tries += 1

    # If 3 attempts fail, return False
    returnResponse['type'] = 'Error'
    returnResponse['message'] = errorMessage if errorMessage != None else 'OTHER'
    return returnResponse

def get_details_url(url):
    tries = 1
    all_data = []
    unique_ids = set()  # To track unique cell IDs
    part_index = 1  # Global part index for all records

    returnResponse = dict()

    while tries <= 3:  # Try a maximum of 3 times
        try:
            print('Details try', tries)
            driver.get(url)
            time.sleep(3)

            # Try to find the scrollable container
            scrollable_div = driver.find_element(By.CSS_SELECTOR, ".ReactVirtualized__Grid.grid")

            # Initial scroll to the top
            driver.execute_script("arguments[0].scrollTop = 0", scrollable_div)  # Scroll to top
            page = 1

            while True:
                # Capture the data currently displayed
                detail_soup = BeautifulSoup(driver.page_source, 'html.parser')
                current_data = extract_data_from_soup(detail_soup, unique_ids, part_index)
                print(f'current data (Page {page}):', current_data)

                # Update part_index to reflect the total number of records processed
                part_index += len(current_data)

                # Add only unique records to all_data
                all_data.extend(current_data)

                # If no new records were found, stop the loop
                if len(current_data) == 0:
                    print('No new data found, stopping scroll.')
                    break

                # Scroll to load more data
                driver.execute_script("arguments[0].scrollTop += arguments[0].clientHeight", scrollable_div)
                time.sleep(1)  # Wait for new content to load

                page += 1

            returnResponse['type'] = 'Success'
            returnResponse['data'] = all_data
            return returnResponse

        except Exception as e:
            print('catch error:', type(e))
            time.sleep(3)  # Wait before retrying

        tries += 1

    # If 3 attempts fail, return False
    returnResponse['type'] = 'Error'
    returnResponse['message'] = 'NoContentFound'
    return returnResponse


def extract_data_from_soup(soup, unique_ids, part_index):
    data = []
    # Extract details and part numbers
    column_headers_div = soup.find('div', {'class': 'columnHeaders'})
    
    if not column_headers_div:
        print('No column headers found in the page source!')
        return []

    headers = []
    for name_desc_div in column_headers_div.find_all('div', {'class': 'name-desc'}):
        # Get 'columnDesc' text if available
        column_desc = name_desc_div.find('div', {'class': 'columnDesc'})
        desc_text = column_desc.text.strip() if column_desc and column_desc.text.strip() else ''

        # Get 'columnName' text
        column_name = name_desc_div.find('div', {'class': 'columnName'}).find('span', {'class': 'value'})
        name_text = column_name.text.strip() if column_name else ''

        # Combine the two, format like "desc (name)" if desc is present, otherwise just name
        if desc_text:
            headers.append(f"{desc_text} ({name_text})")
        else:
            headers.append(name_text)
    
    rows_div = soup.find('div', {'class': 'cells'})
    if not rows_div:
        print('No rows found in the page source!')
        return []
    
    rows = rows_div.find_all('div', {'class': 'psol-comp-TabExCell-themeable'})

    current_part = {}

    for i, cell in enumerate(rows):
        cell_id = cell.get('id')
        if cell_id and cell_id not in unique_ids:
            unique_ids.add(cell_id)  # Track unique cell ID

            header = headers[i % len(headers)]

            # Find the content inside the cell
            cell_content = cell.find('div', {'class': 'cellContent'})

            # Check if the content contains an image
            img_tag = cell_content.find('img') if cell_content else None

            if img_tag:
                # If an image is found, extract the 'src' attribute
                current_part[header] = img_tag['src']
            else:
                # Otherwise, extract the text content
                 current_part[header] = cell_content.text.strip() if cell_content else ""

            # When the row is fully populated, assign the current global index
            if (i + 1) % len(headers) == 0:
                current_part['Index'] = str(part_index)
                data.append(current_part)
                current_part = {}
                part_index += 1

    return data


# Main scraping function
def scrape(eclass_code, prod_links):
    print('Started Process for ->', eclass_code)
    for prod_link in prod_links:
        if not prod_link in all_links:
            all_links.append(prod_link)
            
            # Navigate to the product list page
            url = prod_link
            datafile = os.path.join(datapath, f'3dfindit_{eclass_code}')
            if os.path.isfile(datafile):
                f1 = open(datafile, 'r')
                soup = BeautifulSoup(f1.read(),'html.parser') 
                f1.close()
            else:
                response = get_url(url)
                if response['type'] == 'Error':
                    return response['message']
            
                # Parse the fully loaded product list page with BeautifulSoup
                soup = BeautifulSoup(driver.page_source, 'html.parser')

                # Save the html source to file
                savehtml(datafile, driver.page_source)
    
            try:
                products_list = []

                # Find all product family elements on the page
                product_families = soup.find_all('div', class_='psol-comp-NodesGridCell')

                for family in product_families:
                    product_anchor_tag = family.find('a', class_='gridCardBody')
                    if product_anchor_tag is None:
                        continue  # Skip if no anchor tag found
                    product_link = product_anchor_tag['href']
                    extracted_data = get_details_url(f"https://www.3dfindit.com{product_link}")
                    if extracted_data['type'] == 'Error':
                        continue

                    # Parse the product detail page
                    detail_soup = BeautifulSoup(driver.page_source, 'html.parser')
                    output = {
                        "eclass": eclass_code,
                        "productID": None,
                        "description": None,
                        "standardNumber": None,
                        "company": None,
                        "latestChange": None,
                        "productURL": None,
                        "productImagePicture": None,
                        "partNumbers": extracted_data['data']
                    }

                    # Extract details like standard number, description, company, etc.
                    output['standardNumber'] = detail_soup.find('td', {'class': 'classtable-entry-right'}).text.strip() if detail_soup.find('td', {'class': 'classtable-entry-right'}) else None
                    output['description'] = detail_soup.find_all('td', {'class': 'classtable-entry-right'})[1].text.strip() if len(detail_soup.find_all('td', {'class': 'classtable-entry-right'})) > 1 else None
                    output['company'] = detail_soup.find_all('td', {'class': 'classtable-entry-right'})[2].text.strip() if len(detail_soup.find_all('td', {'class': 'classtable-entry-right'})) > 2 else None
                    output['latestChange'] = detail_soup.find_all('td', {'class': 'classtable-entry-right'})[3].text.strip() if len(detail_soup.find_all('td', {'class': 'classtable-entry-right'})) > 3 else None
                    output['productImagePicture'] = detail_soup.find('img', {'class': 'nodeImage'})['src'] if detail_soup.find('img', {'class': 'nodeImage'}) else None
                    output['productURL'] = detail_soup.find('meta', {'property': 'og:url'})['content'] if detail_soup.find('meta', {'property': 'og:url'}) else None
                    output['productID'] = detail_soup.find('div', {'class': 'title'}).text.strip() if detail_soup.find('div', {'class': 'title'}) else None
                    
                    print('output', output)

                    products_list.append(output)

                    # Save the extracted product details to a JSON file
                    filename = os.path.join(jsonpath, sanitize(f'3dpartfindit-{eclass_code}') + '.json')
                    with open(filename, 'w', encoding='utf-8') as json_file:
                        json.dump(products_list, json_file, ensure_ascii=False, indent=4)
                    

            except Exception as e:
                flog.write(eclass_code + ':' + str(e) + '\n')
                flog.flush()  # Log any exceptions that occur
                return str(e)

# Prompt the user whether to continue from where the script left off or start fresh
yn = input('Continue where it left off ? (y/n): ')
yn = yn.upper()

if yn == 'N':
    files = glob.glob(os.path.join(datapath, '*'))
    for file in files:
        try:
            os.unlink(file)
        except OSError as e:
            print("Error: %s : %s" % (file, e.strerror))  # Handle any errors while deleting files

# Prompt for delay between operations
delay = input('Delay(seconds, 0=no delay): ')
delay = int(delay)

# Open the Excel file containing eclass codes and URLs
columns = ['Product Name', 'Manufacturer', 'Part ID', 'Description', 'Part Number', 'Table']
flog = open('3dfindit.log', 'w', encoding='utf8')
excel_file = 'eclass_codes.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# Start processing each row in the Excel file
line = 2

while not ws[f'B{line}'].value is None:
    if ws[f'C{line}'].value != 'YES':
        start_url = ws[f'B{line}'].value
        code = int(ws[f'E{line}'].value)
        response = scrape(code, [start_url])
        print('message got', response)
        if response == 'NoSuchElementException':
            ws[f'D{line}'].value = 'No records'
            ws[f'C{line}'].value = 'YES'
        elif response == 'OTHER':
            pass
        else:
            ws[f'C{line}'].value = 'YES'
        wb.save(excel_file)  # Mark the row as processed in the Excel file
    line += 1
flog.close()  # Close the log file after processing is complete
